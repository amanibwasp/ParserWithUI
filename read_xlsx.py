import openpyxl
def read_xlsx(path):
    vendor_codes = []
    if path == '':
        return 'empty'
    if path.split('.')[-1] != 'xlsx':
        return 'not xlsx'
    wb = openpyxl.load_workbook(path)
    wh = wb.active
    max_row = wh.max_row
    for row in range(1, max_row + 1):
        vendor_codes.append(wh.cell(row=row, column=1).value)
    return vendor_codes
